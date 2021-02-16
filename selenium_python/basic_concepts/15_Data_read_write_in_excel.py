import openpyxl

path="D:/Users/Mukesh/Desktop/selenium_python/basic_concepts/data_1.xlsx"

workbook=openpyxl.load_workbook(path)
sheet=workbook.active    # workbook.get_sheet_by_name("sheet1")

rows=sheet.max_row
cols=sheet.max_column

print(rows)
print(cols)

 ##--------------reading data from excel sheet-------------------
#for r in range(1,rows+1):
#    for c in range(1,cols+1):
#        print(sheet.cell(row=r,column=c).value,end="       ")
#    print()

 ##--------------writing data to excel sheet-------------------
#for r in range(rows+2,rows+3):
#    for c in range(cols+2,cols+3):
#        sheet.cell(row=r,column=c).value="welcome"
    
#workbook.save(path)
#-------------------------------------------------------------------

